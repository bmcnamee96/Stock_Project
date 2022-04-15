from openpyxl.workbook import Workbook

# headers       = ['Company','Address','Tel','Web']
# workbook_name = 'sample.xlsx'
# wb = Workbook()
# page = wb.active
# page.title = 'companies'
# page.append(headers) # write the headers to the first line

# # Data to write:
# companies = [['name1','address1','tel1','web1'], ['name2','address2','tel2','web2']]

# for info in companies:
#     page.append(info)
# wb.save(filename = workbook_name)

from openpyxl import load_workbook

workbook_name = 'sample.xlsx'
wb = load_workbook(workbook_name)
page = wb.active

# New data to write:
new_companies = [['name3','address3','tel3','web3'], ['name4','address4','tel4','web4']]

for info in new_companies:
    page.append(info)

wb.save(filename=workbook_name)